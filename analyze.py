import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('C:/Users/jennm/Desktop/Dashboard Files/Hospital POC Base Data.xlsx', data_only=True)
ws = wb['Base Data']

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        data.append({
            'facility': row[0],
            'admits': row[1] or 0,
            'payor_summary': row[5],
            'admit_month': row[8],
            'admit_year': row[9],
            'qtr': row[10],
            'program': row[11],
            'program_type': row[12],
            'financial_class': row[14],
            'los': row[16] or 0,
            'gross_charges': row[17] or 0,
            'net_charges': row[18] or 0,
            'denial_adj': row[21] or 0,
            'denied_days': row[31] or 0,
            'auth_days': row[30] or 0,
            'prim_pmt': row[32] or 0,
        })

print(f"Total records: {len(data)}")

# By facility
by_fac = defaultdict(list)
for d in data:
    by_fac[d['facility']].append(d)

print("\n=== Per Facility Totals ===")
for fac in sorted(by_fac):
    recs = by_fac[fac]
    total_los = sum(r['los'] for r in recs)
    total_gross = sum(r['gross_charges'] for r in recs)
    total_net = sum(r['net_charges'] for r in recs)
    total_admits = sum(r['admits'] for r in recs)
    print(f"{fac}: admits={total_admits}, LOS={total_los}, Gross=${total_gross:,.0f}, Net=${total_net:,.0f}")

print("\n=== Financial Class by Facility ===")
for fac in sorted(by_fac):
    recs = by_fac[fac]
    fc_data = defaultdict(lambda: {'los':0,'gross':0,'net':0,'admits':0})
    for r in recs:
        fc = r['financial_class']
        fc_data[fc]['los'] += r['los']
        fc_data[fc]['gross'] += r['gross_charges']
        fc_data[fc]['net'] += r['net_charges']
        fc_data[fc]['admits'] += r['admits']
    total_los = sum(v['los'] for v in fc_data.values())
    print(f"\n{fac}:")
    for fc in sorted(fc_data):
        pct = (fc_data[fc]['los']/total_los*100) if total_los else 0
        rpd = (fc_data[fc]['net']/fc_data[fc]['los']) if fc_data[fc]['los'] else 0
        print(f"  {fc}: LOS={fc_data[fc]['los']}, {pct:.1f}%, Gross=${fc_data[fc]['gross']:,.0f}, Net=${fc_data[fc]['net']:,.0f}, RevPerDay=${rpd:.2f}")

print("\n=== Top 5 Payors by Gross (all qtrs) ===")
for fac in sorted(by_fac):
    recs = by_fac[fac]
    payor_gross = defaultdict(float)
    for r in recs:
        payor_gross[r['payor_summary']] += r['gross_charges']
    top5 = sorted(payor_gross.items(), key=lambda x: -x[1])[:5]
    print(f"\n{fac} - Top 5 by Gross:")
    for p, v in top5:
        print(f"  {p}: ${v:,.0f}")

print("\n=== Rolling 5 Qtrs - Top 5 Payors by Gross (LOS) ===")
qtrs = ['1Q25','2Q25','3Q25','4Q25','1Q26']
for fac in sorted(by_fac):
    recs = by_fac[fac]
    payor_gross = defaultdict(float)
    for r in recs:
        payor_gross[r['payor_summary']] += r['gross_charges']
    top5 = [p for p, v in sorted(payor_gross.items(), key=lambda x: -x[1])[:5]]
    print(f"\n{fac} - Top 5 Gross Payors by LOS per Qtr:")
    for p in top5:
        row_data = []
        for q in qtrs:
            los = sum(r['los'] for r in recs if r['payor_summary']==p and r['qtr']==q)
            row_data.append(f"{q}:{los}")
        print(f"  {p}: {', '.join(row_data)}")

print("\n=== 2026 YTD Denial Rate by Payor ===")
for fac in sorted(by_fac):
    recs = [r for r in by_fac[fac] if r['qtr']=='1Q26']
    payor_data = defaultdict(lambda: {'gross':0,'denial':0})
    for r in recs:
        payor_data[r['payor_summary']]['gross'] += r['gross_charges']
        payor_data[r['payor_summary']]['denial'] += abs(r['denial_adj'])
    print(f"\n{fac} - 2026 YTD Denial Rate:")
    for p in sorted(payor_data):
        g = payor_data[p]['gross']
        d = payor_data[p]['denial']
        rate = (d/g*100) if g else 0
        if g > 0:
            print(f"  {p}: Gross=${g:,.0f}, DenialAdj=${d:,.0f}, Rate={rate:.1f}%")

# Covered Lives
ws_cl = wb['Covered Lives']
print("\n=== Covered Lives ===")
for row in ws_cl.iter_rows(min_row=1, max_row=30, values_only=True):
    print(row)

# Regional Info
ws_ri = wb['Hospital Regional Info']
print("\n=== Regional Info (all rows) ===")
for row in ws_ri.iter_rows(min_row=1, values_only=True):
    print(row)
