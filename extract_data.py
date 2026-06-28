import openpyxl
import json
from collections import defaultdict

wb = openpyxl.load_workbook('C:/Users/jennm/Desktop/Dashboard Files/Hospital POC Base Data.xlsx', data_only=True)
ws = wb['Base Data']

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        data.append({
            'facility': row[0],
            'admits': row[1] or 0,
            'payor_summary': row[5],
            'admit_month': row[8],
            'admit_year': str(row[9]) if row[9] else '',
            'qtr': row[10],
            'program_type': row[12],
            'financial_class': row[14],
            'los': row[16] or 0,
            'gross_charges': row[17] or 0,
            'net_charges': row[18] or 0,
            'denial_adj': row[21] or 0,
            'denied_days': row[31] or 0,
            'auth_days': row[30] or 0,
        })

QTRS = ['1Q25','2Q25','3Q25','4Q25','1Q26']
FACILITIES = sorted(set(d['facility'] for d in data))

output = {}

for fac in FACILITIES:
    recs = [d for d in data if d['facility'] == fac]

    # 1. Payor Mix (Financial Class % by LOS)
    fc_data = defaultdict(lambda: {'los':0,'gross':0,'net':0,'admits':0})
    for r in recs:
        fc = r['financial_class']
        fc_data[fc]['los'] += r['los']
        fc_data[fc]['gross'] += r['gross_charges']
        fc_data[fc]['net'] += r['net_charges']
        fc_data[fc]['admits'] += r['admits']
    total_los = sum(v['los'] for v in fc_data.values())

    payor_mix = {}
    for fc, vals in fc_data.items():
        rpd = (vals['net'] / vals['los']) if vals['los'] else 0
        pct = (vals['los'] / total_los * 100) if total_los else 0
        avg_los = (vals['los'] / vals['admits']) if vals['admits'] else 0
        payor_mix[fc] = {
            'los': round(vals['los']),
            'gross': round(vals['gross'], 2),
            'net': round(vals['net'], 2),
            'admits': round(vals['admits']),
            'pct': round(pct, 1),
            'rev_per_day': round(rpd, 2),
            'avg_los': round(avg_los, 1)
        }

    # 2. Top 5 by Gross Charges - rolling LOS by quarter
    payor_gross = defaultdict(float)
    for r in recs:
        payor_gross[r['payor_summary']] += r['gross_charges']
    top5_gross_payors = [p for p, _ in sorted(payor_gross.items(), key=lambda x: -x[1])[:5]]

    top5_gross = {}
    for p in top5_gross_payors:
        qtr_los = {}
        for q in QTRS:
            los = sum(r['los'] for r in recs if r['payor_summary'] == p and r['qtr'] == q)
            qtr_los[q] = los
        total_gross = round(payor_gross[p], 2)
        top5_gross[p] = {'qtr_los': qtr_los, 'total_gross': total_gross}

    # 3. Top 5 by Net Charges - rolling LOS by quarter
    payor_net = defaultdict(float)
    for r in recs:
        payor_net[r['payor_summary']] += r['net_charges']
    top5_net_payors = [p for p, _ in sorted(payor_net.items(), key=lambda x: -x[1])[:5]]

    top5_net = {}
    for p in top5_net_payors:
        qtr_los = {}
        for q in QTRS:
            los = sum(r['los'] for r in recs if r['payor_summary'] == p and r['qtr'] == q)
            qtr_los[q] = los
        total_net = round(payor_net[p], 2)
        top5_net[p] = {'qtr_los': qtr_los, 'total_net': total_net}

    # 4. 2026 YTD Denial Rate
    recs_2026 = [r for r in recs if r['qtr'] == '1Q26']
    payor_denial = defaultdict(lambda: {'gross':0,'net':0,'denial':0,'los':0,'admits':0})
    for r in recs_2026:
        p = r['payor_summary']
        payor_denial[p]['gross'] += r['gross_charges']
        payor_denial[p]['net'] += r['net_charges']
        payor_denial[p]['denial'] += abs(r['denial_adj'])
        payor_denial[p]['los'] += r['los']
        payor_denial[p]['admits'] += r['admits']

    denial_rate = {}
    for p, vals in payor_denial.items():
        rate = (vals['denial'] / vals['gross'] * 100) if vals['gross'] else 0
        denial_rate[p] = {
            'gross': round(vals['gross'], 2),
            'net': round(vals['net'], 2),
            'denial_adj': round(vals['denial'], 2),
            'rate': round(rate, 2),
            'los': round(vals['los']),
            'admits': round(vals['admits'])
        }

    # 5. Quarterly summary (all payors, all FC)
    qtr_summary = {}
    for q in QTRS:
        qrecs = [r for r in recs if r['qtr'] == q]
        qtr_summary[q] = {
            'los': sum(r['los'] for r in qrecs),
            'gross': round(sum(r['gross_charges'] for r in qrecs), 2),
            'net': round(sum(r['net_charges'] for r in qrecs), 2),
            'admits': sum(r['admits'] for r in qrecs),
        }

    # 6. Program type breakdown
    pt_data = defaultdict(lambda: {'los':0,'gross':0,'net':0,'admits':0})
    for r in recs:
        pt = r['program_type'] or 'Unknown'
        pt_data[pt]['los'] += r['los']
        pt_data[pt]['gross'] += r['gross_charges']
        pt_data[pt]['net'] += r['net_charges']
        pt_data[pt]['admits'] += r['admits']
    program_mix = {pt: {'los': v['los'], 'gross': round(v['gross'],2), 'net': round(v['net'],2), 'admits': v['admits']} for pt, v in pt_data.items()}

    output[fac] = {
        'payor_mix': payor_mix,
        'top5_gross': top5_gross,
        'top5_gross_payors': top5_gross_payors,
        'top5_net': top5_net,
        'top5_net_payors': top5_net_payors,
        'denial_rate': denial_rate,
        'qtr_summary': qtr_summary,
        'program_mix': program_mix,
        'totals': {
            'los': sum(r['los'] for r in recs),
            'gross': round(sum(r['gross_charges'] for r in recs), 2),
            'net': round(sum(r['net_charges'] for r in recs), 2),
            'admits': sum(r['admits'] for r in recs),
        }
    }

# Covered Lives
ws_cl = wb['Covered Lives']
covered_lives_raw = []
for row in ws_cl.iter_rows(min_row=2, values_only=True):
    if row[1]:  # Row Labels
        covered_lives_raw.append({
            'hospital': row[0] or 'N/A',
            'payor': row[1],
            'total': row[2] or 0,
            'medicare_advantage': row[3] or 0,
            'commercial': row[4] or 0,
            'managed_medicaid': row[5] or 0,
        })

# Regional Info
ws_ri = wb['Hospital Regional Info']
regional_info = []
for row in ws_ri.iter_rows(min_row=2, values_only=True):
    if row[1]:
        regional_info.append({
            'id': row[0],
            'name': row[1],
            'location': row[2],
            'region': row[3],
            'vp_finance': row[4],
            'vp_operations': row[5],
            'market_ceo': row[6],
            'ceo': row[7],
            'cfo': row[8],
        })

final = {
    'hospitals': output,
    'covered_lives': covered_lives_raw,
    'regional_info': regional_info,
    'quarters': QTRS,
    'facilities': FACILITIES,
}

with open('C:/Users/jennm/Desktop/Dashboard Files/dashboard_data.json', 'w') as f:
    json.dump(final, f, indent=2)

print("Done! dashboard_data.json created.")
print(f"Facilities: {FACILITIES}")
