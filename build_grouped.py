import openpyxl, json
from collections import defaultdict

wb = openpyxl.load_workbook('Hospital POC Base Data.xlsx', read_only=True)
ws = wb['Base Data']
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
idx = {h: i for i, h in enumerate(hdr) if h}

groups = defaultdict(lambda: {'los': 0, 'gross': 0.0, 'net': 0.0, 'denial_adj': 0.0, 'admits': 0})

for row in ws.iter_rows(min_row=2, values_only=True):
    fac = row[idx['Facility Name']]
    if fac is None:
        continue
    qtr = row[idx['QTR']]
    ptype = row[idx['Program Type']] or 'Unknown'
    fc = row[idx['Financial Class']]
    payor = row[idx['Payor Summary']]
    los = row[idx['LOS']] or 0
    gross = row[idx['Total Gross Charges']] or 0
    contr = row[idx['Contractual Adjustments']] or 0
    net = gross + contr
    denial = abs(row[idx['Denial Adjustments']] or 0)
    admits = row[idx['Admits']] or 0
    key = (fac, qtr, ptype, fc, payor)
    g = groups[key]
    g['los'] += los
    g['gross'] += gross
    g['net'] += net
    g['denial_adj'] += denial
    g['admits'] += admits

records = []
for (fac, qtr, ptype, fc, payor), v in groups.items():
    records.append({
        'f': fac, 'q': qtr, 'pt': ptype, 'fc': fc, 'p': payor,
        'los': v['los'], 'g': round(v['gross'], 2), 'n': round(v['net'], 2),
        'd': round(v['denial_adj'], 2), 'a': v['admits']
    })

with open('grouped_records.json', 'w') as f:
    json.dump(records, f, separators=(',', ':'))

print('groups:', len(groups))
print('wrote', len(records), 'records')
print('sample', records[:2])
